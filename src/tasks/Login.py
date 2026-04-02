import json
import os
import re
import time
from datetime import datetime, timedelta

import openpyxl
import pyautogui
import pyperclip
import requests

from src.tasks.MyBaseTask import MyBaseTask

MAIL_TM_API = "https://api.mail.tm"
ACCOUNT_PATH = r"C:\Users\chris\Documents\AIWork\BLSR\daily_ok_sr\Account.xlsx"
LOGOUT_LOG_PATH = r"C:\Users\chris\Documents\AIWork\BLSR\daily_ok_sr\logout_times.json"


def _load_logout_times():
    if os.path.exists(LOGOUT_LOG_PATH):
        with open(LOGOUT_LOG_PATH, 'r') as f:
            return json.load(f)
    return {}


def _save_logout_time(email):
    data = _load_logout_times()
    data[email] = datetime.now().isoformat()
    with open(LOGOUT_LOG_PATH, 'w') as f:
        json.dump(data, f, indent=2)


def _save_to_xlsx(email, money):
    """Save logout date (col C) and money (col D) to the account's row in Account.xlsx."""
    wb = openpyxl.load_workbook(ACCOUNT_PATH)
    ws = wb.active
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row[0].value == email:
            ws.cell(row=row_idx, column=3).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=4).value = money
            break
    wb.save(ACCOUNT_PATH)


def _next_reset_time():
    """Returns the next 8am reset time."""
    now = datetime.now()
    reset = now.replace(hour=8, minute=0, second=0, microsecond=0)
    if now >= reset:
        reset += timedelta(days=1)
    return reset


def _should_skip(email):
    """Returns True if the account logged out after the last 8am reset."""
    data = _load_logout_times()
    if email not in data or not data[email]:
        return False
    logout_time = datetime.fromisoformat(data[email])
    now = datetime.now()
    last_reset = now.replace(hour=8, minute=0, second=0, microsecond=0)
    if now < last_reset:
        last_reset -= timedelta(days=1)
    return logout_time >= last_reset


def _load_accounts():
    wb = openpyxl.load_workbook(ACCOUNT_PATH)
    ws = wb.active
    return [(row[0], row[1]) for row in ws.iter_rows(values_only=True) if row[0]]


def _get_mail_tm_token(email, password):
    resp = requests.post(f"{MAIL_TM_API}/token", json={
        "address": email,
        "password": password
    }, timeout=10)
    resp.raise_for_status()
    return resp.json()["token"]


def _get_latest_subject(token):
    resp = requests.get(f"{MAIL_TM_API}/messages", headers={
        "Authorization": f"Bearer {token}"
    }, timeout=10)
    resp.raise_for_status()
    members = resp.json().get("hydra:member", [])
    if not members:
        return None
    return members[0]["subject"]


def _extract_code(subject):
    if not subject:
        return None
    match = re.search(r'\d{4,8}', subject)
    return match.group() if match else None


class Login(MyBaseTask):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.name = "登入挖礦賣礦"
        self.description = "自動輪流登入所有帳號，挖礦並賣礦"
        self._accounts = None

    def click_real(self, box):
        """Real foreground click using pyautogui — converts window-relative to screen coords."""
        from ok import og
        win = og.device_manager.capture_method.hwnd_window
        x = int(win.x + box.x + box.width / 2)
        y = int(win.y + box.y + box.height / 2)
        pyautogui.click(x, y)

    def type_real(self, text):
        """Type text via clipboard paste — handles special chars like @ reliably."""
        pyperclip.copy(text)
        pyautogui.hotkey('ctrl', 'v')

    def _fetch_verify_code(self, token, timeout=300):
        try:
            before = _get_latest_subject(token)
        except Exception:
            before = None

        deadline = time.time() + timeout
        while time.time() < deadline:
            self.sleep(5)
            try:
                subject = _get_latest_subject(token)
            except Exception:
                continue
            if subject and subject != before:
                code = _extract_code(subject)
                if code:
                    self.log_info(f"收到驗證碼: {code}")
                    return code
        raise Exception("5分鐘內未收到驗證碼郵件")

    def run(self):
        if self._accounts is None:
            self._accounts = _load_accounts()
            self.log_info(f"載入 {len(self._accounts)} 個帳號")
        if not self._accounts:
            raise Exception("Account.xlsx 中沒有帳號")

        for index, (email, password) in enumerate(self._accounts):
            self.log_info(f"========== 帳號 {index + 1}/{len(self._accounts)}: {email} ==========", notify=True)
            if _should_skip(email):
                self.log_info(f"帳號 {email} 今日已完成，跳過", notify=True)
                continue
            self._login(email, password)
            self.dig_ores()
            self.sell_ores(email)
            self._logout(email)

        self.log_info("所有帳號完成！", notify=True)

    # ------------------------------------------------------------------ #
    #  Login                                                               #
    # ------------------------------------------------------------------ #

    def _login(self, email, password):
        # Detect HAOPLAY login interface
        self.log_info("等待 HAOPLAY 登入畫面...")
        self.wait_until(lambda: self.find_one("haoplay_login_screen"), time_out=60)

        # Click email login button
        self.log_info("選擇 Email 登入...")
        email_login_btn = self.wait_until(lambda: self.find_one("email_login_button"), time_out=10)
        self.click_real(email_login_btn)

        # Detect email login interface
        self.log_info("等待 Email 登入介面...")
        email_screen = self.wait_until(lambda: self.find_one("email_login_screen"), time_out=30)
        if not email_screen:
            self.screenshot("email_login_screen_not_found")
            raise Exception("找不到 Email 登入介面，已截圖至 screenshots 資料夾")

        # Connect to mail.tm account
        self.log_info("連接 mail.tm 帳號...")
        token = _get_mail_tm_token(email, password)

        # Click email input field and type email
        self.log_info(f"輸入 Email: {email}")
        self.click_real(email_screen)
        self.sleep(0.3)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('delete')
        self.sleep(0.1)
        self.type_real(email)

        # Click send verification code button
        self.log_info("點擊獲取驗證碼...")
        send_code_btn = self.wait_until(lambda: self.find_one("send_verify_code_button"), time_out=10)
        self.click_real(send_code_btn)

        # Get verification code from email
        self.log_info("等待驗證碼郵件...")
        code = self._fetch_verify_code(token, timeout=300)

        # Enter verification code
        self.log_info(f"輸入驗證碼: {code}")
        code_input = self.wait_until(lambda: self.find_one("verify_code_input"), time_out=10)
        self.click_real(code_input)
        self.sleep(0.3)
        self.type_real(code)

        # Click 登入
        self.log_info("點擊登入...")
        login_btn = self.wait_until(lambda: self.find_one("login_button"), time_out=10)
        self.click_real(login_btn)

        # Click 鏈接開始
        self.log_info("等待鏈接開始...")
        connect_btn = self.wait_until(lambda: self.find_one("connect_start_button"), time_out=60)
        self.click_real(connect_btn)

        # Click 選擇角色開始
        self.log_info("等待選擇角色開始...")
        select_char_btn = self.wait_until(lambda: self.find_one("select_character_button"), time_out=60)
        self.sleep(0.5)
        self.click_real(select_char_btn)

        self.log_info("登入完成！", notify=True)
        self.sleep(2)

    # ------------------------------------------------------------------ #
    #  Dig ores                                                            #
    # ------------------------------------------------------------------ #

    def _get_button_color(self):
        """Sample the pixel color at the digging button location in the game frame."""
        frame = self.frame
        if frame is None:
            return None
        # TODO: replace 960, 600 with the actual pixel coords of the digging button
        x, y = 1460, 561
        return frame[y, x]  # returns [B, G, R]

    def _is_button_red(self):
        color = self._get_button_color()
        if color is None:
            return False
        g, r = int(color[1]), int(color[2])
        return r > 150 and g < 100

    def _is_button_normal(self):
        color = self._get_button_color()
        if color is None:
            return False
        r = int(color[2])
        return r < 150

    def dig_ores(self):
        # Check for beginner popup within 20 seconds and close it
        self.log_info("檢查新手彈窗...")
        popup = self.wait_until(lambda: self.find_one("beginner_popup"), time_out=20)
        if popup:
            self.log_info("發現新手彈窗，關閉中...")
            close_btn = self.wait_until(lambda: self.find_one("close_popup_button"), time_out=10)
            if close_btn:
                self.click_box(close_btn)
                self.sleep(1)

        # Check if stamina already depleted on entry
        if self._is_button_red():
            self.log_info("體力已耗盡，跳過挖礦", notify=True)
            return

        # Dig loop — keep pressing F until red appears
        self.log_info("開始挖礦！", notify=True)
        while True:
            if self._is_button_red():
                self.log_info("體力耗盡，停止挖礦", notify=True)
                break
            self.send_key('f')
            self.sleep(1)

    # ------------------------------------------------------------------ #
    #  Sell ores                                                           #
    # ------------------------------------------------------------------ #

    def sell_ores(self, email=None):
        # Step 1: Press ESC to open menu
        self.log_info("開啟選單...")
        self.send_key('esc')
        self.sleep(1)

        # Step 2: Click trading center button
        self.log_info("進入交易中心...")
        trade_btn = self.wait_until(lambda: self.find_one("trade_center_button"), time_out=15)
        if not trade_btn:
            self.log_error("找不到交易中心按鈕")
            return
        self.click_box(trade_btn)
        self.sleep(2)

        # Step 3: Click 我要賣 button (skip to Step 4 if not found)
        self.log_info("點擊我要賣...")
        sell_tab = self.wait_until(lambda: self.find_one("sell_tab_button"), time_out=10)
        if sell_tab:
            self.click_box(sell_tab)
            self.sleep(1)
        else:
            self.log_info("找不到我要賣按鈕，跳至收回/領取步驟")

        # Step 4: Collect unsold ores or claim money from yesterday
        collect_btn = self.find_one("collect_button")
        if collect_btn:
            self.log_info("發現未售出礦石，點擊收回...")
            self.click_box(collect_btn)
            self.sleep(1)
        else:
            gain_money_btn = self.find_one("gain_money_button")
            if gain_money_btn:
                self.log_info("領取昨日收益...")
                self.click_box(gain_money_btn)
                self.sleep(1)

        # Step 5: Click the ore item to sell — skip to ESC if no ore found
        self.log_info("選擇礦石...")
        ore_item = self.wait_until(lambda: self.find_one("ore_item"), time_out=5)
        if not ore_item:
            self.log_info("沒有礦石可賣，記錄金錢後跳至登出")
            if email:
                money_boxes = self.ocr(0.746354167, 0.0324074, 0.80885416666666666666666666666667, 0.06018518518518518518518518518519, log=False)
                money = next((b.name for b in money_boxes if b.name), None)
                self.log_info(f"當前金錢: {money}")
                _save_to_xlsx(email, money)
            self.send_key('esc')
            self.sleep(1)
            return
        self.click_box(ore_item)
        self.sleep(1)

        # Step 6: Click sell all button
        self.log_info("點擊全部賣出...")
        sell_all_btn = self.wait_until(lambda: self.find_one("sell_all_button"), time_out=5)
        if not sell_all_btn:
            self.log_error("找不到全部賣出按鈕")
            return
        self.click_box(sell_all_btn)
        self.sleep(0.5)

        # Step 7: Adjust price to match the lowest price
        self.log_info("調整售價至最低價...")
        for _ in range(500):
            current_price = self.ocr(0.73, 0.59, 0.76, 0.62, log=False)
            lowest_price  = self.ocr(0.76, 0.62, 0.80, 0.66, log=False)

            if not current_price or not lowest_price:
                self.log_error("OCR 讀取失敗，停止調整")
                break

            cur = next((int(m.group()) for b in current_price
                        if b.name and (m := re.search(r'\d+', str(b.name)))), None)
            low = next((int(m.group()) for b in lowest_price
                        if b.name and (m := re.search(r'\d+', str(b.name)))), None)

            if cur is None or low is None:
                break

            self.log_info(f"當前價格: {cur}  |  最低市場價: {low}")

            if cur == low:
                self.log_info("價格已達最低市場價")
                break
            elif cur < low:
                minus_btn = self.find_one("price_minus_button", use_gray_scale=True)
                if minus_btn:
                    self.click_box(minus_btn, down_time=0.1)
                else:
                    self.log_error("找不到 - 按鈕，停止調整")
                    break
            else:
                plus_btn = self.find_one("price_plus_button", use_gray_scale=True)
                if plus_btn:
                    self.click_box(plus_btn, down_time=0.1)
                else:
                    self.log_error("找不到 + 按鈕，停止調整")
                    break

            self.sleep(0.1)

        # Step 8: Click sell button
        self.log_info("點擊上架...")
        sell_btn = self.wait_until(lambda: self.find_one("list_sell_button"), time_out=5)
        if sell_btn:
            self.click_box(sell_btn)
            self.sleep(0.5)

        # Step 9: Click confirm button
        confirm_btn = self.wait_until(lambda: self.find_one("confirm_button"), time_out=5)
        if confirm_btn:
            self.click_box(confirm_btn)
            self.sleep(0.5)

        # Record money after confirming sell — money is visible in trade center
        if email:
            money_boxes = self.ocr(0.746354167, 0.0324074, 0.80885416666666666666666666666667, 0.06018518518518518518518518518519, log=False)
            money = next((b.name for b in money_boxes if b.name), None)
            self.log_info(f"當前金錢: {money}")
            _save_to_xlsx(email, money)
            self.log_info(f"已記錄 {email} 金錢: {money}")

        # Step 10: Press ESC to close
        self.send_key('esc')
        self.sleep(1)
        self.log_info("賣礦完成！", notify=True)

    # ------------------------------------------------------------------ #
    #  Logout                                                              #
    # ------------------------------------------------------------------ #

    def _logout(self, email=None):
        # Step 11: Click logout button
        self.log_info("登出中...")
        logout_btn = self.wait_until(lambda: self.find_one("logout_button"), time_out=15)
        if not logout_btn:
            self.log_error("找不到登出按鈕")
            return
        self.click_real(logout_btn)
        self.sleep(1)

        # Step 12: Click confirm logout button if it appears
        confirm_logout_btn = self.wait_until(lambda: self.find_one("confirm_logout_button"), time_out=20)
        if confirm_logout_btn:
            self.click_real(confirm_logout_btn)
            self.sleep(1)

        # Step 13: Click switch/logout account button
        switch_btn = self.wait_until(lambda: self.find_one("switch_account_button"), time_out=10)
        if not switch_btn:
            self.log_error("找不到更換帳號按鈕")
            return
        self.click_real(switch_btn)
        self.sleep(2)

        if email:
            _save_logout_time(email)
            self.log_info(f"已記錄 {email} 登出時間")
        self.log_info("已登出", notify=True)
